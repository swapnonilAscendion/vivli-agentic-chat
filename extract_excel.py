import sys
import io
import openpyxl
from openpyxl import load_workbook

# Fix Unicode output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

excel_path = r"C:\Users\swapnonil.mukherjee\projects\vivli-chatbot\Vivli_AI_Chat_User_Stories(1).xlsx"

wb = load_workbook(excel_path)

print("=" * 120)
print("USER STORIES - VIVLI AI CHAT")
print("=" * 120)
print()

for sheet_name in wb.sheetnames:
    print(f"\n{'='*120}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*120}\n")

    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=False):
        for cell in row:
            if cell.value:
                print(f"{cell.value}")
        print()
